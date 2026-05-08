from flask import Flask, render_template, request, jsonify, send_file
import requests
import io
import os
import re
import time
from difflib import SequenceMatcher
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

load_dotenv()

app = Flask(__name__)

OSU_CLIENT_ID = os.getenv("OSU_CLIENT_ID")
OSU_CLIENT_SECRET = os.getenv("OSU_CLIENT_SECRET")

TOKEN_URL = "https://osu.ppy.sh/oauth/token"
API_BASE = "https://osu.ppy.sh/api/v2"

_token_cache = {
    "access_token": None,
    "expires_at": 0,
}

MOD_ROW_COLORS = {
    "DT+HR": "DDA0DD",
    "DT":    "B0E0E6",
    "HR":    "FFB6C1",
    "HD":    "FFFACD",
    "FM":    "FFE4B5",
    "TB":    "DDA0DD",
    "NM":    "F0F0F0",
}


def get_osu_token():
    if not OSU_CLIENT_ID or not OSU_CLIENT_SECRET:
        raise RuntimeError("Missing OSU_CLIENT_ID or OSU_CLIENT_SECRET in .env")

    now = time.time()

    if _token_cache["access_token"] and now < _token_cache["expires_at"]:
        return _token_cache["access_token"]

    response = requests.post(
        TOKEN_URL,
        json={
            "client_id": int(OSU_CLIENT_ID),
            "client_secret": OSU_CLIENT_SECRET,
            "grant_type": "client_credentials",
            "scope": "public",
        },
        timeout=20,
    )

    response.raise_for_status()
    data = response.json()

    _token_cache["access_token"] = data["access_token"]
    _token_cache["expires_at"] = now + data.get("expires_in", 86400) - 60

    return _token_cache["access_token"]


def osu_request(method, path, params=None, json_body=None):
    token = get_osu_token()

    response = requests.request(
        method,
        f"{API_BASE}{path}",
        headers={
            "Authorization": f"Bearer {token}",
            "Accept": "application/json",
            "Content-Type": "application/json",
        },
        params=params or {},
        json=json_body,
        timeout=20,
    )

    response.raise_for_status()
    return response.json()


def get_beatmap(map_id):
    return osu_request("GET", f"/beatmaps/{map_id}")


def get_beatmap_attributes(map_id, mods=None, ruleset="osu"):
    body = {
        "ruleset": ruleset,
    }

    if mods:
        body["mods"] = mods

    data = osu_request(
        "POST",
        f"/beatmaps/{map_id}/attributes",
        json_body=body,
    )

    return data.get("attributes", {})


def ar_to_ms(ar):
    return 1800 - 120 * ar if ar <= 5 else 1200 - 150 * (ar - 5)


def ms_to_ar(ms):
    return (1800 - ms) / 120 if ms >= 1200 else 5 + (1200 - ms) / 150


def apply_mods(bm, active_mods):
    cs = float(bm.get("cs", 0) or 0)
    ar = float(bm.get("ar", 0) or 0)
    od = float(bm.get("accuracy", 0) or 0)
    hp = float(bm.get("drain", 0) or 0)
    bpm = float(bm.get("bpm", 0) or 0)

    if "HR" in active_mods:
        cs = min(cs * 1.3, 10)
        ar = min(ar * 1.4, 10)
        od = min(od * 1.4, 10)
        hp = min(hp * 1.4, 10)

    if "DT" in active_mods:
        bpm *= 1.5
        ar = ms_to_ar(ar_to_ms(ar) / 1.5)
        od = (79.5 - (79.5 - 6 * od) / 1.5) / 6
        ar = min(max(ar, 0), 11)
        od = min(max(od, 0), 11)

    return {
        "cs": round(cs, 2),
        "ar": round(ar, 2),
        "od": round(od, 2),
        "hp": round(hp, 2),
        "bpm": round(bpm, 1),
    }


def row_fill_color(active_mods, slot_type=""):
    if slot_type == "TB":
        return MOD_ROW_COLORS["TB"]

    no_fm = [m for m in active_mods if m != "FM"]

    if "DT" in no_fm and "HR" in no_fm:
        key = "DT+HR"
    elif no_fm:
        key = no_fm[0]
    elif "FM" in active_mods:
        key = "FM"
    else:
        key = "NM"

    return MOD_ROW_COLORS.get(key, "FFFFFF")


def star_fill_color(sr):
    try:
        sr = float(sr)
    except (ValueError, TypeError):
        return "FFFFFF"

    if sr < 2:
        return "BBBBBB"
    if sr < 3:
        return "88CC66"
    if sr < 4:
        return "DDDD44"
    if sr < 5:
        return "FF9922"
    if sr < 6:
        return "FF6644"
    if sr < 7:
        return "EE44AA"

    return "AA22CC"


def mod_string(mods):
    return "+".join(mods) if mods else "NM"


def safe_float(value, default=""):
    try:
        return f"{float(value):.2f}"
    except (ValueError, TypeError):
        return default


def normalize_text(value):
    value = str(value or "").lower()
    value = re.sub(r"[^a-z0-9]+", " ", value)
    value = re.sub(r"\s+", " ", value).strip()
    return value


def similarity(a, b):
    a = normalize_text(a)
    b = normalize_text(b)

    if not a or not b:
        return 0

    return SequenceMatcher(None, a, b).ratio()


def parse_length_to_seconds(value):
    if not value:
        return None

    match = re.match(r"^(\d+):(\d{1,2})$", str(value).strip())

    if not match:
        return None

    minutes = int(match.group(1))
    seconds = int(match.group(2))

    return minutes * 60 + seconds


def score_closeness(actual, expected, max_points, scale):
    if actual is None or expected is None:
        return 0

    try:
        actual = float(actual)
        expected = float(expected)
    except (ValueError, TypeError):
        return 0

    diff = abs(actual - expected)
    return max(0, max_points - diff * scale)


def slot_to_mods(slot):
    slot = str(slot or "").upper()

    if slot.startswith("HD"):
        return ["HD"]
    if slot.startswith("HR"):
        return ["HR"]
    if slot.startswith("DT"):
        return ["DT"]
    if slot.startswith("FM"):
        return ["FM"]

    return []


def slot_to_type(slot):
    slot = str(slot or "").upper()

    if slot.startswith("HD"):
        return "HD"
    if slot.startswith("HR"):
        return "HR"
    if slot.startswith("DT"):
        return "DT"
    if slot.startswith("FM"):
        return "FM"
    if slot.startswith("TB"):
        return "TB"

    return "NM"


def clean_search_query(title, version=""):
    raw = f"{title} {version}".strip()
    cleaned = re.sub(r"[^A-Za-z0-9\s]", " ", raw)
    cleaned = re.sub(r"\s+", " ", cleaned).strip()

    return cleaned or raw


def search_beatmapsets(query):
    if not query:
        return []

    try:
        data = osu_request(
            "GET",
            "/beatmapsets/search",
            params={
                "q": query,
                "m": 0,
            },
        )
    except requests.RequestException:
        return []

    return data.get("beatmapsets", []) or []


def candidate_from_beatmapset(beatmapset, beatmap):
    return {
        "beatmap_id": str(beatmap.get("id", "")),
        "beatmapset_id": str(beatmapset.get("id", beatmap.get("beatmapset_id", ""))),
        "title": beatmapset.get("title", ""),
        "artist": beatmapset.get("artist", ""),
        "version": beatmap.get("version", ""),
        "stars": beatmap.get("difficulty_rating", beatmap.get("difficulty_rating", None)),
        "bpm": beatmap.get("bpm", beatmapset.get("bpm", None)),
        "ar": beatmap.get("ar", None),
        "od": beatmap.get("accuracy", None),
        "hp": beatmap.get("drain", None),
        "cs": beatmap.get("cs", None),
        "length_seconds": beatmap.get("total_length", beatmap.get("hit_length", None)),
        "raw": beatmap,
    }


def get_search_candidates(slot):
    title = slot.get("title", "")
    version = slot.get("version", "")

    queries = [
        clean_search_query(title, version),
        clean_search_query(title),
    ]

    seen_queries = []
    candidates = []
    seen_ids = set()

    for query in queries:
        if not query or query in seen_queries:
            continue

        seen_queries.append(query)
        beatmapsets = search_beatmapsets(query)

        for beatmapset in beatmapsets[:12]:
            for beatmap in beatmapset.get("beatmaps", []) or []:
                beatmap_id = str(beatmap.get("id", ""))

                if not beatmap_id or beatmap_id in seen_ids:
                    continue

                seen_ids.add(beatmap_id)
                candidates.append(candidate_from_beatmapset(beatmapset, beatmap))

    return candidates


def score_candidate(slot, candidate):
    slot_mods = slot.get("mods", [])
    raw_bm = candidate.get("raw", {})

    adjusted = apply_mods(raw_bm, slot_mods)
    expected_length = parse_length_to_seconds(slot.get("length"))

    actual_length = candidate.get("length_seconds")
    if actual_length is not None and "DT" in slot_mods:
        actual_length = float(actual_length) / 1.5

    score = 0

    title_score = similarity(slot.get("title", ""), candidate.get("title", "")) * 25
    version_score = similarity(slot.get("version", ""), candidate.get("version", "")) * 35

    score += title_score
    score += version_score

    score += score_closeness(adjusted.get("bpm"), slot.get("bpm"), 10, 0.25)
    score += score_closeness(adjusted.get("ar"), slot.get("ar"), 5, 4)
    score += score_closeness(adjusted.get("od"), slot.get("od"), 5, 4)
    score += score_closeness(adjusted.get("hp"), slot.get("hp"), 3, 2)
    score += score_closeness(adjusted.get("cs"), slot.get("cs"), 3, 2)
    score += score_closeness(actual_length, expected_length, 7, 0.35)

    try:
        attrs = get_beatmap_attributes(candidate["beatmap_id"], mods=[m for m in slot_mods if m != "FM"])
        attr_star = attrs.get("star_rating")
        score += score_closeness(attr_star, slot.get("stars"), 12, 8)
        candidate["resolved_star"] = attr_star
    except requests.RequestException:
        candidate["resolved_star"] = candidate.get("stars")
        score += score_closeness(candidate.get("stars"), slot.get("stars"), 6, 4)

    confidence = max(0, min(100, round(score)))

    candidate["confidence"] = confidence
    candidate["adjusted_bpm"] = adjusted.get("bpm")
    candidate["adjusted_ar"] = adjusted.get("ar")
    candidate["adjusted_od"] = adjusted.get("od")
    candidate["adjusted_hp"] = adjusted.get("hp")
    candidate["adjusted_cs"] = adjusted.get("cs")

    return candidate


def resolve_romai_slot(slot):
    candidates = get_search_candidates(slot)

    scored = []
    for candidate in candidates[:20]:
        if not candidate.get("beatmap_id"):
            continue

        scored.append(score_candidate(slot, candidate))

    scored.sort(key=lambda item: item.get("confidence", 0), reverse=True)

    best = scored[0] if scored else None

    resolved = dict(slot)

    if best:
        resolved["beatmap_id"] = best.get("beatmap_id")
        resolved["beatmapset_id"] = best.get("beatmapset_id")
        resolved["artist"] = best.get("artist", "")
        resolved["matched_title"] = best.get("title", "")
        resolved["matched_version"] = best.get("version", "")
        resolved["confidence"] = best.get("confidence", 0)
        resolved["match_status"] = "matched" if best.get("confidence", 0) >= 70 else "review"
        resolved["candidates"] = [
            {
                "beatmap_id": c.get("beatmap_id"),
                "beatmapset_id": c.get("beatmapset_id"),
                "title": c.get("title"),
                "artist": c.get("artist"),
                "version": c.get("version"),
                "confidence": c.get("confidence"),
                "resolved_star": c.get("resolved_star"),
                "adjusted_bpm": c.get("adjusted_bpm"),
            }
            for c in scored[:3]
        ]
    else:
        resolved["confidence"] = 0
        resolved["match_status"] = "not_found"
        resolved["candidates"] = []

    return resolved


def parse_romai_pool_text(raw_text):
    text = raw_text.replace("\r\n", "\n").replace("\r", "\n")
    lines = [line.strip() for line in text.split("\n") if line.strip()]

    pool = {
        "name": "",
        "average_stars": None,
        "elo": None,
        "slots": [],
    }

    slot_re = re.compile(r"^(NM|HD|HR|DT|FM)\d+$|^TB\d*$", re.IGNORECASE)
    info_re = re.compile(
        r"^(?P<title>.+)\s+\[(?P<version>[^\]]+)\]\s+"
        r"(?P<stars>\d+(?:\.\d+)?)★\s+BPM:\s*(?P<bpm>\d+(?:\.\d+)?)",
        re.IGNORECASE,
    )
    stats_re = re.compile(
        r"Stats:\s*AR:(?P<ar>\d+(?:\.\d+)?)\s*\|\s*"
        r"OD:(?P<od>\d+(?:\.\d+)?)\s*\|\s*"
        r"HP:(?P<hp>\d+(?:\.\d+)?)\s*\|\s*"
        r"CS:(?P<cs>\d+(?:\.\d+)?)\s*--\s*"
        r"Length:\s*(?P<length>\d+:\d{1,2})",
        re.IGNORECASE,
    )

    for i, line in enumerate(lines):
        if line.lower().startswith("detected pool"):
            if not pool["name"] and i + 1 < len(lines):
                next_line = lines[i + 1].strip()
                if not next_line.lower().startswith("average stars"):
                    pool["name"] = next_line
            continue

        if line.lower().startswith("average stars"):
            avg_match = re.search(r"Average Stars:\s*(\d+(?:\.\d+)?)★", line, re.IGNORECASE)
            elo_match = re.search(r"ELO:\s*(\d+)", line, re.IGNORECASE)

            if avg_match:
                pool["average_stars"] = float(avg_match.group(1))

            if elo_match:
                pool["elo"] = int(elo_match.group(1))

            continue

    idx = 0

    while idx < len(lines):
        line = lines[idx]

        if not slot_re.match(line):
            idx += 1
            continue

        slot = line.upper()

        if idx + 1 >= len(lines):
            idx += 1
            continue

        map_line = lines[idx + 1]
        info_match = info_re.match(map_line)

        if not info_match:
            idx += 1
            continue

        slot_data = {
            "slot": slot,
            "slot_type": slot_to_type(slot),
            "title": info_match.group("title").strip(),
            "version": info_match.group("version").strip(),
            "stars": float(info_match.group("stars")),
            "bpm": float(info_match.group("bpm")),
            "ar": None,
            "od": None,
            "hp": None,
            "cs": None,
            "length": None,
            "beatmap_id": None,
            "mods": slot_to_mods(slot),
            "confidence": None,
            "match_status": "unresolved",
        }

        if idx + 2 < len(lines):
            stats_line = lines[idx + 2]
            stats_match = stats_re.search(stats_line)

            if stats_match:
                slot_data["ar"] = float(stats_match.group("ar"))
                slot_data["od"] = float(stats_match.group("od"))
                slot_data["hp"] = float(stats_match.group("hp"))
                slot_data["cs"] = float(stats_match.group("cs"))
                slot_data["length"] = stats_match.group("length")

        pool["slots"].append(slot_data)
        idx += 3

    if not pool["name"]:
        pool["name"] = "Imported RomAI Pool"

    return pool


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/load_maps", methods=["POST"])
def load_maps():
    raw_ids = request.json.get("beatmaps", "")
    ids = [x for x in raw_ids.replace(",", " ").split() if x.isdigit()]

    maps = []

    for mid in ids:
        try:
            bm = get_beatmap(mid)
        except requests.RequestException:
            continue

        beatmapset = bm.get("beatmapset") or {}

        maps.append({
            "beatmap_id": str(bm.get("id", "")),
            "beatmapset_id": str(bm.get("beatmapset_id", "")),
            "title": beatmapset.get("title", ""),
            "artist": beatmapset.get("artist", ""),
            "version": bm.get("version", ""),
        })

    return jsonify(maps)


@app.route("/beatmap_lookup", methods=["POST"])
def beatmap_lookup():
    beatmap_id = str(request.json.get("beatmap_id", "")).strip()

    if not beatmap_id.isdigit():
        return jsonify({
            "error": "Beatmap ID must be numeric."
        }), 400

    try:
        bm = get_beatmap(beatmap_id)
    except requests.RequestException:
        return jsonify({
            "error": "Beatmap could not be found or is unavailable."
        }), 404

    beatmapset = bm.get("beatmapset") or {}

    return jsonify({
        "beatmap_id": str(bm.get("id", "")),
        "beatmapset_id": str(bm.get("beatmapset_id", "")),
        "title": beatmapset.get("title", ""),
        "artist": beatmapset.get("artist", ""),
        "version": bm.get("version", ""),
        "stars": bm.get("difficulty_rating", None),
        "bpm": bm.get("bpm", None),
        "ar": bm.get("ar", None),
        "od": bm.get("accuracy", None),
        "hp": bm.get("drain", None),
        "cs": bm.get("cs", None),
        "length_seconds": bm.get("total_length", bm.get("hit_length", None)),
    })


@app.route("/parse_romai_pool", methods=["POST"])
def parse_romai_pool():
    raw_text = request.json.get("text", "")

    if not raw_text.strip():
        return jsonify({
            "error": "No RomAI output provided."
        }), 400

    parsed = parse_romai_pool_text(raw_text)

    if not parsed["slots"]:
        return jsonify({
            "error": "No map slots found. Make sure you pasted the /mappools show output."
        }), 400

    return jsonify(parsed)


@app.route("/resolve_romai_pool", methods=["POST"])
def resolve_romai_pool():
    pool = request.json.get("pool")

    if not pool or not pool.get("slots"):
        return jsonify({
            "error": "No parsed RomAI pool provided."
        }), 400

    resolved_slots = []

    for slot in pool["slots"]:
        resolved_slots.append(resolve_romai_slot(slot))

    resolved_pool = dict(pool)
    resolved_pool["slots"] = resolved_slots

    matched = len([slot for slot in resolved_slots if slot.get("match_status") == "matched"])
    review = len([slot for slot in resolved_slots if slot.get("match_status") == "review"])
    not_found = len([slot for slot in resolved_slots if slot.get("match_status") == "not_found"])

    resolved_pool["resolve_summary"] = {
        "matched": matched,
        "review": review,
        "not_found": not_found,
        "total": len(resolved_slots),
    }

    return jsonify(resolved_pool)


@app.route("/generate", methods=["POST"])
def generate():
    maps = request.json.get("maps", [])

    wb = Workbook()
    ws = wb.active
    ws.title = "Beatmaps"

    headers = [
        "Map ID", "Title", "Artist", "Version",
        "Slot", "★ Star", "BPM", "HP Drain",
        "Max Combo", "CS", "AR", "OD", "Thumbnail"
    ]

    ws.append(headers)

    hdr_fill = PatternFill("solid", fgColor="1E1E2E")
    hdr_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = hdr_align

    ws.row_dimensions[1].height = 24

    col_widths = [14, 30, 22, 22, 12, 8, 10, 10, 12, 8, 8, 8, 22]

    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    for m in maps:
        map_id = m.get("beatmap_id")
        active_mods = m.get("mods", [])
        slot_type = m.get("slot_type", "")
        slot_label = m.get("slot_label", "")

        if not map_id:
            continue

        try:
            bm = get_beatmap(map_id)
        except requests.RequestException:
            continue

        beatmapset = bm.get("beatmapset") or {}

        api_mods = [mod for mod in active_mods if mod != "FM"]

        try:
            attrs = get_beatmap_attributes(map_id, mods=api_mods)
        except requests.RequestException:
            attrs = {}

        star_val = attrs.get("star_rating", bm.get("difficulty_rating", ""))
        star_str = safe_float(star_val)

        max_combo = attrs.get("max_combo", bm.get("max_combo", ""))

        adj = apply_mods(bm, active_mods)

        beatmap_id = bm.get("id")
        beatmapset_id = bm.get("beatmapset_id")

        map_link = f'=HYPERLINK("https://osu.ppy.sh/beatmaps/{beatmap_id}","{beatmap_id}")'
        thumbnail = f'=IMAGE("https://b.ppy.sh/thumb/{beatmapset_id}.jpg",4,120,160)'

        row_data = [
            map_link,
            beatmapset.get("title", ""),
            beatmapset.get("artist", ""),
            bm.get("version", ""),
            slot_label or slot_type or mod_string(active_mods),
            star_str,
            adj["bpm"],
            adj["hp"],
            max_combo,
            adj["cs"],
            adj["ar"],
            adj["od"],
            thumbnail,
        ]

        ws.append(row_data)

        xlsx_row = ws.max_row
        row_fill = PatternFill("solid", fgColor=row_fill_color(active_mods, slot_type))
        center = Alignment(horizontal="center", vertical="center")

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=xlsx_row, column=col_idx)
            cell.fill = row_fill
            cell.alignment = center
            cell.font = Font(name="Arial", size=10)

        if star_str:
            star_cell = ws.cell(row=xlsx_row, column=6)
            star_cell.fill = PatternFill("solid", fgColor=star_fill_color(star_str))
            star_cell.font = Font(bold=True, name="Arial", size=10)

        ws.row_dimensions[xlsx_row].height = 20

    mem = io.BytesIO()
    wb.save(mem)
    mem.seek(0)

    return send_file(
        mem,
        as_attachment=True,
        download_name="beatmaps.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
